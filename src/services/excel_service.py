from typing import (
    TypeAlias,
    Union,
    Any
)
from datetime import date, datetime
from pathlib import Path
import os

from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell


AnyType: TypeAlias = Union[str, int, float, date, datetime]


class ExcelService:
    @staticmethod
    def load_file(path: str | Path, **kwargs) -> Workbook:
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        return load_workbook(path, **kwargs)


    @staticmethod
    def create_file(**kwargs) -> Workbook:
        wb = Workbook(**kwargs)

        return wb
    
    @staticmethod
    def data_to_cells(data: list[Any], ws: Worksheet) -> list[Cell]:
        return [
            Cell(worksheet=ws, value=el) for el in data
        ]
    

    @staticmethod
    def cells_to_worksheet(rows: list[list[Cell]], ws: Worksheet) -> None:
        for row in rows:
            ws.append(row)
    

    @staticmethod
    def read_worksheet_by_row(ws: Worksheet, **kwargs) -> list[list[AnyType]]:
        return [
            list(row) for row in ws.iter_rows(**kwargs)
        ]
    
    @staticmethod
    def read_worksheet_by_column(ws: Worksheet, **kwargs) -> list[list[AnyType]]:
        return [
            list(column) for column in ws.iter_cols(**kwargs)
        ]
    

    @staticmethod
    def fit_worksheet(ws: Worksheet) -> None:
        for e in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(e)].bestFit = True
            ws.column_dimensions[get_column_letter(e)].auto_size = True


    @staticmethod
    def get_line_chart(ws: Workbook, reference: Reference, **kwargs) -> LineChart:
        
        chart = LineChart()

        chart.title = kwargs.get("chart_title", "default chart name")

        chart.x_axis.crossAx = kwargs.get("x_crossAx", 100)
        chart.y_axis.crossAx = kwargs.get("y_crossAx", 500)

        chart.x_axis.title = kwargs.get("x_axis_title", "default name")

        chart.y_axis.title = kwargs.get("y_axis_title", "default name")

        chart.width = kwargs.get("chart_width", 15)
            
        chart.height = kwargs.get("chart_height", 7.5)
        
        data = Reference(ws, **kwargs)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(reference)
        
        match kwargs.get("style"):
            case "default":
                chart.style = 12
            case _:
                chart.style = 12

        return chart
    
    @staticmethod
    def add_chart(ws: Worksheet, chart, coordinate: str) -> None:
        ws.add_chart(chart, coordinate)


    @staticmethod
    def style_like_header_rows(rows: list[list[Cell]]) -> list[list[Cell]]:
        styled_rows = []

        for row in rows:
            styled_row = []

            for cell in row:
                styled_row.append(
                    ExcelService.set_header_cell_style(cell)
                )
            
            styled_rows.append(styled_row)
        
        return styled_rows


    @staticmethod
    def style_like_body_rows(rows: list[list[Cell]]) -> list[list[Cell]]:
        styled_rows = []

        for e, row in enumerate(rows):
            styled_row = []

            for cell in row:
                styled_row.append(
                    ExcelService.set_body_cell_style(cell, e)
                )
            
            styled_rows.append(styled_row)
        
        return styled_rows

        
    @staticmethod
    def set_header_cell_style(cell: Cell) -> Cell:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='003366FF', end_color='003366FF', fill_type='solid')
        cell.border = Border(
            right = Side(color="FF000000", style="thin")
        )

        return cell

    
    @staticmethod
    def set_body_cell_style(cell: Cell, e: int) -> Cell:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            right = Side(color="FF000000", style="thin")
        )

        if e % 2 == 1:
            cell.fill = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type='solid')

        if e % 2 == 0:
            cell.fill = PatternFill(start_color='0033CCCC', end_color='0033CCCC', fill_type='solid')

        return cell
