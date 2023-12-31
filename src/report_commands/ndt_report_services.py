"""
NDT report command can dump ndt data as excel file and pdf file
Excel file will contain the main sheet (Main) and other sheets named after the kleymo
PDF file will contain the first page (Main) with general data and other pages with ndt data belonging to welders
"""

from dataclasses import dataclass
from itertools import chain
from datetime import date
from copy import copy
from typing import (
    Literal, 
    TypeAlias,
    Union
)
import re

from openpyxl.styles import Alignment, PatternFill, NamedStyle, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart.axis import DateAxis
from openpyxl.cell.cell import Cell

from src.domain import NDTRequest, DBResponse, WelderNDTModel
from src.services.utils import reverse_dict, limit_dict_pair, limit_dict_values
from settings import SEARCH_VALUES_FILE, STATIC_DIR, NDT_REPORT_PATH
from src.db.repository import NDTRepository
from src.services.utils import load_json


"""
=======================================================================================================
Types
=======================================================================================================
"""


Saver: TypeAlias = Union["NDTReportExcelSaveService", "NDTReportPDFSaveService"]
SortedRows: TypeAlias = dict[str, list['DataRow']]
ExcelRow: TypeAlias = list[Cell]
Any: TypeAlias = str | int | float | date


"""
=======================================================================================================
Value Objects
=======================================================================================================
"""


@dataclass
class MainPageData:
    summarized_welder_ndts: list["DataRow"]
    summarized_welder_group_ndts: dict[date, "DataRow"]


@dataclass
class DataRow:
    full_name: str | None = None
    kleymo: str | int = None
    sicil_number: str | None = None
    birthday: str | None = None
    passport_number: str | int | None = None
    nation: str | None = None
    comp: str | None = None
    subcon: str | None = None
    project: str | None = None
    latest_welding_date: str = None
    total_weld_1: float | None = None
    total_ndt_1: float | None = None
    total_accepted_1: float | None = None
    total_repair_1: float | None = None
    repair_status_1: float | None = None
    total_weld_2: float | None = None
    total_ndt_2: float | None = None
    total_accepted_2: float | None = None
    total_repair_2: float | None = None
    repair_status_2: float | None = None
    total_weld_3: float | None = None
    total_ndt_3: float | None = None
    total_accepted_3: float | None = None
    total_repair_3: float | None = None
    repair_status_3: float | None = None
    ndt_id: str = None

    def to_row(self, ws: Worksheet) -> ExcelRow:
        row: ExcelRow = []

        for el in self.__dict__.values():
            cell = Cell(ws)

            if el == None:
                el = 0

            cell.value = el
            row.append(cell)

        return row
    

    def to_main_sheet_row(self, ws: Worksheet) -> ExcelRow:
        row = self.to_row(ws)

        return [row[1]] + row[10:-1]


"""
=======================================================================================================
Domain Services
=======================================================================================================
"""


class NDTReportService:

    def report(self, search_date: str | None, save_mode: str, limit: int | None) -> None:
        ndts = RequestNDTsService(search_date=search_date).request_data().result
        preprocessor = NDTDataPrepocessor()

        saver = self._get_saver(save_mode)
        preprocessor.preprocess(ndts=ndts, limit=limit)


        saver.dump_report(
            *preprocessor.preprocess(ndts=ndts, limit=limit)
        )


    def _get_saver(self, save_mode: str) -> Saver:
        
        match save_mode:
            case 'excel':
                return NDTReportExcelSaveService()
            case 'pdf':
                return NDTReportPDFSaveService()
            case _:
                raise ValueError("Invalid save_mode")
    

class NDTDataPrepocessor:

    def preprocess(self, ndts: list[WelderNDTModel], limit: str | None) -> tuple[MainPageData, SortedRows]:

        rows = [self._remove_none_values(DataRow(**ndt.__dict__)) for ndt in ndts]

        limit = limit if limit != None else 10

        sorted_rows_by_kleymo = self._sort_by_kleymo(rows)
        sorted_rows_by_date = limit_dict_pair(self._sort_by_date(sorted_rows_by_kleymo), limit)

        main_page_data = MainPageData(
            summarized_welder_ndts = [self._get_cumulated_row(row_list) for row_list in limit_dict_values(sorted_rows_by_kleymo, limit).values()],
            summarized_welder_group_ndts = self._summarize_group(reverse_dict(sorted_rows_by_date))
        )

        return (main_page_data, limit_dict_values(sorted_rows_by_kleymo, limit))


    def _remove_none_values(self, row: DataRow) -> DataRow:
        row = list(row.__dict__.items())
        result = []

        for el in row[10:-1]:
            if el[1] ==None:
                result.append(
                    (el[0], 0)
                )
                continue

            result.append(el)
        
        return DataRow(**dict(row[:10] + result))
    

    def _summarize_group(self, rows_dict: SortedRows) -> dict[str, DataRow]:
        return {
            key: self._compute_repair_statuses(self._sum_rows(value)) for key, value in rows_dict.items()
        }

    
    def _sum_rows(self, rows: list[DataRow]) -> DataRow:
        return DataRow(
            total_weld_1 = sum([row.total_weld_1 for row in rows]),
            total_ndt_1 = sum([row.total_ndt_1 for row in rows]),
            total_accepted_1 = sum([row.total_accepted_1 for row in rows]),
            total_repair_1 = sum([row.total_repair_1 for row in rows]),
            total_weld_2 = sum([row.total_weld_2 for row in rows]),
            total_ndt_2 = sum([row.total_ndt_2 for row in rows]),
            total_accepted_2 = sum([row.total_accepted_2 for row in rows]),
            total_repair_2 = sum([row.total_repair_2 for row in rows]),
            total_weld_3 = sum([row.total_weld_3 for row in rows]),
            total_ndt_3 = sum([row.total_ndt_3 for row in rows]),
            total_accepted_3 = sum([row.total_accepted_3 for row in rows]),
            total_repair_3 = sum([row.total_repair_3 for row in rows]),
        )  
    
 
    def _sort_by_kleymo(self, rows: list[DataRow]) -> SortedRows:     
        sorted_rows: dict[str, list[DataRow]] = {}

        for ndt in rows:
            kleymo = ndt.kleymo

            ndt = self._compute_repair_statuses(ndt)

            ndt = DataRow(
                **ndt.__dict__
            )

            if kleymo not in sorted_rows:
                sorted_rows[kleymo] = [ndt]
                continue

            sorted_rows[kleymo].append(ndt)
        
        return sorted_rows


    def _sort_by_date(self, rows: SortedRows) -> dict[date, list[DataRow]]:
        sorted_rows: dict[date, list[DataRow]] = {}
        rows: list[DataRow] = list(chain.from_iterable(rows.values()))

        for row in rows:
            welding_date = row.latest_welding_date

            if welding_date not in sorted_rows:
                sorted_rows[welding_date] = [row]
                continue

            sorted_rows[welding_date].append(row)
        
        return sorted_rows


    def _get_cumulated_row(self, rows: list[DataRow]) -> DataRow:
        rows = list(reversed(rows))

        cumulated_row = rows[0]

        for row in rows[1:]:
            cumulated_row = self._compute_ndt_by_previous(row, cumulated_row)
        
        return cumulated_row
    

    def _compute_ndt_by_previous(self, target_ndt: DataRow, previous_ndt: DataRow) -> DataRow:
        result = []

        target_ndt = list(target_ndt.__dict__.items())
        previous_ndt = list(previous_ndt.__dict__.items())

        for target_el, previous_el in zip(target_ndt[10:-1], previous_ndt[10:-1]):

            if target_el[1] < previous_el[1]:
                result.append(
                    (target_el[0], target_el[1] + previous_el[1])
                )
                continue

            result.append(target_el)
        
        return DataRow(**dict(target_ndt[:10] + result))


    def _compute_repair_statuses(self, row: DataRow) -> DataRow:
        
        row.repair_status_1 = self._compute_repair_status(row.total_accepted_1, row.total_ndt_1)
        row.repair_status_2 = self._compute_repair_status(row.total_accepted_2, row.total_ndt_2)
        row.repair_status_3 = self._compute_repair_status(row.total_accepted_3, row.total_ndt_3)

        return row
    

    def _compute_repair_status(self, total_accepted: float, total_ndt: float):
        try:
            return 100 - (total_accepted / total_ndt) * 100
        except:
            return 0.0


"""
=======================================================================================================
Infrastructure Services
=======================================================================================================
"""


class RequestNDTsService:
    repository = NDTRepository()
    """
    This Services creates NDTRequest object for NDTRepository
    """
    def __init__(self, search_date: str | None) -> None:
        self.search_date = search_date


    def _detect_search_key_type(self, search_key: str) -> Literal["name", "kleymo", "certification_number"]:
        if re.fullmatch(r"[\w]+-[\w]+-[I]+-[0-9]+", search_key.strip()):
            return "certification_number"
        
        elif re.fullmatch(r"[A-Z0-9]{4}", search_key.strip()):
            return "kleymo"
        
        return "name"
        

    def _get_search_values(self, search_values: list[str]) -> tuple[list[str] | None, list[str | int] | None, list[str]] | None:
        names = []
        kleymos = []
        certification_numbers = []

        if len(search_values) == 0:
            return (None, None, None)

        for search_value in search_values:

            match self._detect_search_key_type(search_value):
                case "name":
                    names.append(search_value)
                case "kleymo":
                    kleymos.append(search_value)
                case "certification_number":
                    certification_numbers.append(search_value)

        return (names, kleymos, certification_numbers)
    

    def _get_kleymos_by_search_date(self) -> list[str | int]:
        res = self.repository.get(
            NDTRequest(
                date_before = self.search_date,
                date_from = self.search_date,
            )
        )

        return [ndt.kleymo for ndt in res.result]
    

    def _get_names_kleymos_certification_numbers(self, setting_dict: dict[str, str | list]) -> tuple[list[str], list[str | int], list[str]]:

        if self.search_date != None:
            return (None, self._get_kleymos_by_search_date(), None)
        
        return self._get_search_values(setting_dict.get("search_values", []))
    
    
    def _dump_ndt_request(self) -> NDTRequest:
        ndt_search_settings = load_json(SEARCH_VALUES_FILE)["ndt"]

        names, kleymos, certification_numbers = self._get_names_kleymos_certification_numbers(ndt_search_settings)

        return NDTRequest(
            names = names,
            kleymos = kleymos,
            certification_numbers = certification_numbers,
            comps = ndt_search_settings["comps"],
            subcomps = ndt_search_settings["subcomps"],
            projects = ndt_search_settings["projects"],
            date_from = ndt_search_settings["date_from"],
            date_before = ndt_search_settings["date_before"],
        )
    

    def request_data(self) -> DBResponse[WelderNDTModel]:
        return self.repository.get_many(
            self._dump_ndt_request()
        )


"""
=======================================================================================================
Infrastructure Services
NDT Report excel preprocessor
=======================================================================================================
"""


class NDTReportExcelPreprocessor:
    """
    This class prepare ndt report excel file for further filling by data
    """
    def __init__(self) -> None:
        self.wb = load_workbook(NDT_REPORT_PATH)
    
    
    def _truncate_main_sheet(self) -> None:
        ws: Worksheet = self.wb["Main"]
        ws._charts = []

        ws.delete_rows(3, ws.max_row)


    def _delete_welder_sheets(self) -> None:
        sheets = self.wb.get_sheet_names()

        for sheet in sheets:
            if sheet != "Main":
                self.wb.remove_sheet(self.wb[sheet])


    def _reg_header_style(self) -> None:
        header_style = NamedStyle("header_style")

        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.fill = PatternFill(start_color='003366FF', end_color='003366FF', fill_type='solid')

        self.wb.add_named_style(header_style)


    def get_workbook(self) -> Workbook:
        self._truncate_main_sheet()
        self._delete_welder_sheets()
        try:
            self._reg_header_style()
        except:
            pass

        return self.wb


"""
=======================================================================================================
Infrastructure Services
Report savers
=======================================================================================================
"""
 

class NDTReportExcelSaveService:

    def dump_report(self, main_sheet_data: MainPageData, sorted_ndts: SortedRows) -> None:
        self.wb = NDTReportExcelPreprocessor().get_workbook()
        self._create_welder_sheets(sorted_ndts)
        self._create_main_sheet(main_sheet_data)

        self.wb.save(f"{STATIC_DIR}/report.xlsx")

        # self._set_hyper_links()


    def _create_main_sheet(self, data: MainPageData) -> None:
        ws: Worksheet = self.wb.get_sheet_by_name("Main")

        for e, row in enumerate(data.summarized_welder_ndts):
            row = row.to_main_sheet_row(ws)

            row = self._style_main_sheet(row, e)

            ws.append(row)
        
        summarized_group_rows = []

        for row_date, row in data.summarized_welder_group_ndts.items():
            date_cell = Cell(ws)
            date_cell.value = row_date.strftime("%d.%m.%Y")

            summarized_group_rows.append([date_cell] + row.to_main_sheet_row(ws)[1:])
        
        summarized_group_rows = self._set_coordinates(summarized_group_rows, start_column=18, start_row=3)

        for e, row in enumerate(summarized_group_rows):
            row = self._style_main_sheet(row, e)
            for cell in row:
                ws[cell.coordinate].value = cell.value
                ws[cell.coordinate].style = cell.style
                ws[cell.coordinate].fill = copy(cell.fill)
                ws[cell.coordinate].border = copy(cell.border)
        
        dates = Reference(ws, min_col=18, min_row=3, max_row=2 + len(summarized_group_rows))

        chart1 = self._create_chart(ws, "NDT for Quantity welded pipe joints", min_col=20, max_col=22, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="numbers")
        chart2 = self._create_chart(ws, "NDT for Length of pipe weld joint", min_col=25, max_col=27, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="mm")
        chart3 = self._create_chart(ws, "Structural NDT", min_col=30, max_col=32, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="mm")
        chart4 = self._create_chart(ws, "NDT for Quantity welded pipe joints Repair status", min_col=23, max_col=23, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="percents")
        chart5 = self._create_chart(ws, "NDT for Length of pipe weld joint Repair status", min_col=28, max_col=28, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="percents")
        chart6 = self._create_chart(ws, "Structural NDT Repair status", min_col=33, max_col=33, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="percents")
        chart7 = self._create_chart(ws, "NDT for Quantity welded pipe joints total weld", min_col=19, max_col=19, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="numbers")
        chart8 = self._create_chart(ws, "NDT for Length of pipe weld joint total weld", min_col=24, max_col=24, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="mm")
        chart9 = self._create_chart(ws, "Structural NDT total weld", min_col=29, max_col=29, min_row=2, max_row=2 + len(summarized_group_rows), dates=dates, x_axis="dates", y_axis="mm")

        ws.add_chart(chart1, "S15")
        ws.add_chart(chart2, "X15")
        ws.add_chart(chart3, "AC15")
        ws.add_chart(chart4, "S25")
        ws.add_chart(chart5, "X25")
        ws.add_chart(chart6, "AC25")
        ws.add_chart(chart7, "S35")
        ws.add_chart(chart8, "X35")
        ws.add_chart(chart9, "AC35")


    def _set_coordinates(self, rows: list[ExcelRow], start_column: int, start_row: int) -> list[ExcelRow]:
        result = []

        for row_index, row in enumerate(rows):
            result_row = []

            for column_index, cell in enumerate(row):
                cell.column = start_column + column_index
                cell.row = start_row + row_index

                result_row.append(cell)
            
            result.append(result_row)
    
        return result

    
    def _create_welder_sheets(self, ndts: SortedRows) -> None:
        for kleymo, row_list in ndts.items():
            ws: Worksheet = self.wb.create_sheet(kleymo)

            self._set_table_header_row(
                list(row_list[0].__dict__.keys()),
                ws
            )

            for e, row in enumerate(row_list):
                row = row.to_row(ws)

                row = self._style_body_row(row, e)

                ws.append(row)

            
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].bestFit = True
                ws.column_dimensions[get_column_letter(i)].auto_size = True
            
            self._add_charts(ws)

    
    def _set_table_header_row(self, data: list[str | float | None], ws: Worksheet) -> None:
        row: ExcelRow = []

        for el in data:
            cell = Cell(ws)

            cell.value = el
            cell.style = 'header_style'
            row.append(cell)
        
        ws.append(row)

    
    def _set_hyper_links(self) -> None:
        wb = load_workbook(NDT_REPORT_PATH)
        ws = wb.get_sheet_by_name("Main")

        for e, row in enumerate(ws.iter_rows(min_row=3)):
            cell = row[0]

            cell.hyperlink = f"#{cell.value.strip()}!A1"
            cell.style = "Hyperlink"
            if e % 2 == 1:
                cell.fill = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type='solid')

            if e % 2 == 0:
                cell.fill = PatternFill(start_color='0033CCCC', end_color='0033CCCC', fill_type='solid')

            cell.border = Border(
                right = Side(color="FF000000", style="thick"),
                bottom = Side(color = 'FF000000', style="thin"),
                top = Side(color = 'FF000000', style="thin")
            )

        wb.save(NDT_REPORT_PATH)

    
    def _style_body_row(self, row: ExcelRow, e: int) -> ExcelRow:
        styled_row = []

        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if e % 2 == 1:
                cell.fill = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type='solid')

            if e % 2 == 0:
                cell.fill = PatternFill(start_color='0033CCCC', end_color='0033CCCC', fill_type='solid')
            
            cell.border = Border(
                bottom = Side(color = 'FF000000', style="thin"),
                top = Side(color = 'FF000000', style="thin")
            )

            styled_row.append(cell)
        
        return styled_row
    

    def _style_main_sheet(self, row: ExcelRow, e: int) -> ExcelRow:

        right_bold_border = Border(
            right = Side(color="FF000000", style="thick"),
            bottom = Side(color = 'FF000000', style="thin"),
            top = Side(color = 'FF000000', style="thin")
        )

        row = self._style_body_row(row, e)

        row[-6].border = right_bold_border
        row[-11].border = right_bold_border
        row[-16].border = right_bold_border

        red_fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")

        
        if row[-1].value > 5:
            row[-1].fill = red_fill

        if row[-6].value > 5:
            row[-6].fill = red_fill

        if row[-11].value > 5:
            row[-11].fill = red_fill

        return row
            

    def _add_charts(self, ws: Worksheet) -> None:
        dates = Reference(ws, min_col=10, min_row=2, max_row=ws.max_row)

        chart1 = self._create_chart(ws, "NDT for Quantity welded pipe joints", min_col=12, max_col=14, min_row=1, max_row=ws.max_row, dates=dates)
        chart2 = self._create_chart(ws, "NDT for Length of pipe weld joint", min_col=17, max_col=19, min_row=1, max_row=ws.max_row, dates=dates)
        chart3 = self._create_chart(ws, "Structural NDT", min_col=22, max_col=24, min_row=1, max_row=ws.max_row, dates=dates)
        chart4 = self._create_chart(ws, "NDT for Quantity welded pipe joints total weld", min_col=11, max_col=11, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="numbers")
        chart5 = self._create_chart(ws, "NDT for Length of pipe weld joint total weld", min_col=16, max_col=16, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="mm")
        chart6 = self._create_chart(ws, "Structural NDT total weld", min_col=21, max_col=21, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="mm")
        chart7 = self._create_chart(ws, "NDT for Quantity welded pipe joints Repair status", min_col=15, max_col=15, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="percents")
        chart8 = self._create_chart(ws, "NDT for Length of pipe weld joint Repair status", min_col=20, max_col=20, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="percents")
        chart9 = self._create_chart(ws, "Structural NDT Repair status", min_col=25, max_col=25, min_row=1, max_row=ws.max_row, dates=dates, x_axis="dates", y_axis="percents")

        ws.add_chart(chart1, "A15")
        ws.add_chart(chart2, "H15")
        ws.add_chart(chart3, "O15")
        ws.add_chart(chart4, "A25")
        ws.add_chart(chart5, "H25")
        ws.add_chart(chart6, "O25")
        ws.add_chart(chart7, "A35")
        ws.add_chart(chart8, "H35")
        ws.add_chart(chart9, "O35")

    
    def _create_chart(self, ws: Worksheet, title: str, min_col: int, max_col: int, min_row: int, max_row: int, dates: Reference, x_axis: str | None = None, y_axis: str | None = None) -> LineChart:
        chart = LineChart()
        
        chart.title = title
        chart.style = 12
        chart.x_axis = DateAxis(crossAx=100)
        chart.y_axis.crossAx = 500

        if x_axis:
            chart.x_axis.title = x_axis

        if y_axis:
            chart.y_axis.title = y_axis

        chart.x_axis.number_format = 'd-m-yyyy'
        chart.x_axis.majorTimeUnit = "days"
        chart.width = 20
        
        data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)

        return chart


class NDTReportPDFSaveService:

    def dump_report(self) -> None: ...
