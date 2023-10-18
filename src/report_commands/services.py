from typing import Protocol, Literal, Sequence, TypeVar, TypeAlias, Any
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, PatternFill, NamedStyle, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

from src.domain import NDTRequest, DBResponse, NDTModel,  Name, Kleymo, CertificationNumber
from src.db.repository import NDTRepository
from settings import SEARCH_VALUES_FILE, STATIC_DIR


"""
=======================================================================================================
Types
=======================================================================================================
"""


Saver = TypeVar("Saver", bound='IReportSaveService')
SortedNDTs: TypeAlias = dict[str, list[NDTModel]]
Row: TypeAlias = Sequence[Cell]


"""
=======================================================================================================
Domain (Business) Services
=======================================================================================================
"""


class IReportService(Protocol):
    saver: 'IReportSaveService'

    def report(self, save_mode: str) -> None:
        ...


class NDTReportService(IReportService):

    def _set_saver(self, save_mode: str) -> None:
        match save_mode:
            case 'excel':
                self.saver = NDTReportExcelSaveService()
            case 'pdf':
                self.saver = NDTReportPDFSaveService()

    
    def _limit_ndts(self, limit: int | None) -> SortedNDTs:
                
        ndt_dict: dict[str, list[NDTModel]] = {}

        if limit == None:
            limit = 100


        for ndt in RequestNDTsService().request_data().result:
            if ndt.kleymo not in ndt_dict:
                ndt_dict[ndt.kleymo] = [ndt]
                ndt_dict[ndt.kleymo].append(ndt)
                continue

            if len(ndt_dict[ndt.kleymo]) < limit:
                ndt_dict[ndt.kleymo].append(ndt)
                continue
    
        return ndt_dict


    def report(self, save_mode: str, limit: int | None) -> None:
        self._set_saver(save_mode)

        ndts = self._limit_ndts(limit)

        self.saver.dump_report(ndts)

        

"""
=======================================================================================================
Infrastructure Services
=======================================================================================================
"""


class RequestNDTsService:
    repository = NDTRepository("ndt_summary_table")
    """
    This Services creates NDTRequest object for NDTRepository
    """
    def _detect_search_key_type(self, search_key: str) -> Literal["name", "kleymo", "certification_number"]:
        if re.fullmatch(r"[\w]+-[\w]+-[I]+-[0-9]+", search_key.strip()):
            return "certification_number"
        
        elif re.fullmatch(r"[A-Z0-9]{4}", search_key.strip()):
            return "kleymo"
        
        return "name"
    

    def _read_search_values_file(self) -> Sequence[str]:
        with open(SEARCH_VALUES_FILE, "r", encoding='utf-8') as file:
            values = file.readlines()

            file.close()
        
        return [value.strip() for value in values]
    
    
    def _get_search_values(self) -> tuple[Sequence[Name] | None, Sequence[Kleymo] | None, Sequence[CertificationNumber]] | None:
        search_values = self._read_search_values_file()

        names = []
        kleymos = []
        certification_numbers = []

        if len(search_values) == 0:
            return (None, None, None)

        for search_value in search_values:

            match self._detect_search_key_type(search_value):
                case "name":
                    names.append(search_value)
                case "kleymo":
                    kleymos.append(search_value)
                case "certification_number":
                    certification_numbers.append(search_value)

        return (names, kleymos, certification_numbers)
        
    
    def _dump_ndt_request(self) -> NDTRequest:
        names, kleymos, certification_numbers = self._get_search_values()

        return NDTRequest(
            names = names,
            kleymos = kleymos,
            certification_numbers = certification_numbers,
        )
    

    def request_data(self) -> DBResponse[NDTModel]:
        return self.repository.get_by_request(
            self._dump_ndt_request()
        )


class IReportSaveService(Protocol):

    def dump_report(self, row: Sequence[Any]) -> None: ...
    


class NDTReportExcelSaveService(IReportSaveService):


    def dump_report(self, ndts: SortedNDTs) -> None:
        self.wb = Workbook()
        self._reg_header_style()

        for kleymo, values in ndts.items():
            ws: Worksheet = self.wb.create_sheet(kleymo)
            ws.append(
                self._set_table_header_row(
                    list(values[0].model_dump().keys()), ws
                )
            )
            for e, value in enumerate(values):
                ws.append(
                    self._set_tabe_body_row(
                        list(value.model_dump(mode='json').values()), ws, e
                    )
                )
            
            for i in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(i)].bestFit = True
                ws.column_dimensions[get_column_letter(i)].auto_size = True
            
            self._add_charts(ws)


        self.wb.save(f"{STATIC_DIR}/report.xlsx")

    
    def _set_table_header_row(self, data: Sequence[Any], ws: Worksheet) -> Row:
        row: Row = []

        for e, el in enumerate(data):
            cell = Cell(ws)

            cell.value = el
            cell.style = 'header_style'
            row.append(cell)
        
        return row


    def _set_tabe_body_row(self, data: Sequence[Any], ws: Worksheet, e: int) -> Row:
        row: Row = []

        for el in data:
            cell = Cell(ws)
            cell.value = el

            cell.alignment = Alignment(horizontal='center', vertical='center')

            if e % 2 == 1:
                cell.fill = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type='solid')

            if e % 2 == 0:
                cell.fill = PatternFill(start_color='0033CCCC', end_color='0033CCCC', fill_type='solid')
            
            
            cell.border = Border(
                bottom = Side(color = 'FF000000'),
                top = Side(color = 'FF000000')
            )
            
            row.append(cell)

        return row
            
    
    def _reg_header_style(self) -> None:
        header_style = NamedStyle("header_style")

        header_style.alignment = Alignment(horizontal='center', vertical='center')
        header_style.fill = PatternFill(start_color='003366FF', end_color='003366FF', fill_type='solid')

        self.wb.add_named_style(header_style)

    
    def _add_charts(self, ws: Worksheet) -> None:
        dates = Reference(ws, min_col=9, min_row=2, max_row=ws.max_row)

        chart1 = self._create_chart(ws, "NDT for Quantity welded pipe joints", min_col=10, max_col=14, min_row=1, max_row=ws.max_row, dates=dates)
        chart2 = self._create_chart(ws, "NDT for Length of pipe weld joint", min_col=16, max_col=20, min_row=1, max_row=ws.max_row, dates=dates)
        chart3 = self._create_chart(ws, "Structural NDT", min_col=21, max_col=25, min_row=1, max_row=ws.max_row, dates=dates)

        ws.add_chart(chart1, "A18")
        ws.add_chart(chart2, "H18")
        ws.add_chart(chart3, "O18")

    
    def _create_chart(self, ws: Worksheet, title: str, min_col: int, max_col: int, min_row: int, max_row: int, dates: Reference) -> LineChart:
        chart = LineChart()
        
        chart.title = title
        chart.style = 12
        chart.x_axis = DateAxis(crossAx=100)
        chart.y_axis.crossAx = 500
        chart.x_axis.number_format = 'd-m-yyyy'
        chart.x_axis.majorTimeUnit = "days"
        
        data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)

        chart.add_data(data)
        chart.set_categories(dates)

        return chart

        


class NDTReportPDFSaveService(IReportSaveService):

    def dump_report(self) -> None: ...
