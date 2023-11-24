import re


from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from settings import WELDER_REGISTRY_PATH
from src.domain import WelderModel


class WelderRegistryManager:
    
    number = re.compile(r"[0-9]+")
    value_from = re.compile(r"от [0-9,]+")
    value_before= re.compile(r"до [0-9,]+")


    def update(self, welders: list[WelderModel], group_key: str, subgroup: str, group: str) -> None:
        wb = load_workbook(WELDER_REGISTRY_PATH)
        ws: Worksheet = wb.active
        print(ws.max_row)

        for welder in welders:
            ws.append(
                (
                    ws.max_row + 4,
                    welder.kleymo,
                    "",
                    welder.full_name,
                    group_key,
                    subgroup,
                    "Yes",
                    "",
                    group,
                    self._get_methods(welder),
                    self._get_gtd(self._get_gtd_string(welder)),
                    *self._get_thikness(self._get_thikness_string(welder)),
                    *self._get_diameter(self._get_diameter_string(welder)),
                    self._get_materials(self._get_materials_string(welder))
                )
            )
    
        wb.save(WELDER_REGISTRY_PATH)


    def _get_gtd_string(self, welder: WelderModel) -> str:
        gtds = []

        for certificationn in welder.certifications:
            gtds.append(certificationn.gtd)

        return ", ".join(gtds)


    def _get_materials_string(self, welder: WelderModel) -> str:
        materials = ""

        for certification in welder.certifications:
            materials += f", {certification.groups_materials_for_welding}"

        return materials


    def _get_diameter_string(self, welder: WelderModel) -> str:
        diameters = []

        for certification in welder.certifications:
            if certification.details_diameter != None:
                diameters.append(certification.details_diameter)
            
            if certification.outer_diameter != None:
                diameters.append(certification.outer_diameter)
        
        string = " ||| ".join(diameters)

        return string
    

    def _get_thikness_string(self, welder: WelderModel) -> str:
        thiknesses = []

        for certification in welder.certifications:
            if certification.details_thikness != None:
                thiknesses.append(certification.details_thikness)
        
        string = " ||| ".join(thiknesses)

        return string


    def _get_gtd(self, gtd_string: str) -> str:

        result: dict[str, list[int]] = {}

        for gtd in gtd_string.split("),"):
            key, value = gtd.split("(")

            key = self._get_ndt_abbr(key.strip().lower())

            if key not in result:
                result[key] = [int(i) for i in self.number.findall(value)]
                continue

            result[key] = sorted(set(result[key] + [int(i) for i in self.number.findall(value)]))

        return ", ".join(f"{key}({', '.join(map(str, value))})" for key, value in result.items())
    

    def _get_thikness(self, details_thikness_string: str) -> tuple[float, float]:
        details_thikness_string = details_thikness_string.replace("  ", " ")

        from_strings: list[str] = self.value_from.findall(details_thikness_string)
        before_strings: list[str] = self.value_before.findall(details_thikness_string)

        from_values = [float(value.replace(",", ".").replace("от ", "").replace("свыше ", "").replace("Свыше ", "").strip()) for value in from_strings]
        before_values = [float(value.replace(",", ".").replace("до ", "").strip()) for value in before_strings]
        from_value = min(from_values) if from_values != [] else None
        before_value = max(before_values) if before_values != [] else None

        return (from_value, before_value)
    

    def _get_diameter(self, details_diameter_string: str) -> tuple[float, float]:
        details_diameter_string = details_diameter_string.replace("  ", " ").replace("свыше ", "от ").replace("Свыше ", "от ")

        from_strings: list[str] = self.value_from.findall(details_diameter_string)
        before_strings: list[str] = self.value_before.findall(details_diameter_string)

        from_values = [float(value.replace(",", ".").replace("от ", "").replace("свыше ", "").replace("Свыше ", "").strip()) for value in from_strings]
        before_values = [float(value.replace(",", ".").replace("до ", "").strip()) for value in before_strings]
        from_value = min(from_values) if from_values != [] else None
        before_value = max(before_values) if before_values != [] else None

        return (from_value, before_value)
    

    def _get_materials(self, materials_string: str) -> str:
        materials_string = materials_string.replace("М", "M")

        materials_string = re.sub(r"[[(][\w\W]+[])]", " ", materials_string)

        materials = re.findall(r"[M0-9+]+", materials_string)
        return ", ".join(set(materials))


    def _get_methods(self, welder: WelderModel) -> str:
        methods = []

        for certification in welder.certifications:
            methods.append(certification.method)

        return "/".join(set(methods))


    def _get_ndt_abbr(self, gtd_name: str) -> str:
        if "нефтеперераб" in gtd_name and "взрывопожа" in gtd_name:
            return "ОХНВП"
        if "котел" in gtd_name and "оборуд" in gtd_name:
            return "КО"
        if "нефтегазо" in gtd_name and "оборуд" in gtd_name:
            return "НГДО"
        if "газ" in gtd_name and "оборуд" in gtd_name:
            return "ГО"
        if "строит" in gtd_name and "констру" in gtd_name:
            return "СК"
        if "горнодоб" in gtd_name and "оборуд" in gtd_name:
            return "ГДО"
        if "металлур" in gtd_name and "оборуд" in gtd_name:
            return "МО"
        if "подъемно-транспортное" in gtd_name and "оборуд" in gtd_name:
            return "ПТО"
        if "сталь" in gtd_name and "мост" in gtd_name:
            return "КСМ"
        if "трансп" in gtd_name and "груз" in gtd_name:
            return "ОТОГ"
